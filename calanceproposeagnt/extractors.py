import io
from docx import Document
import fitz  # For handling PDF files
import openpyxl  # For handling .xlsx files

def extract_text_from_file(file):
    file_extension = file.name.split('.')[-1].lower()
    
    if file_extension == 'txt':
        return file.read().decode('utf-8')
    elif file_extension == 'docx':
        doc = Document(io.BytesIO(file.read()))
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])
    elif file_extension == 'pdf':
        pdf_document = fitz.open(stream=file.read(), filetype="pdf")
        text = ""
        for page_num in range(len(pdf_document)):
            page = pdf_document.load_page(page_num)
            text += page.get_text()
        return text
    elif file_extension == 'xlsx':
        workbook = openpyxl.load_workbook(file)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                text += "\t".join(str(cell) for cell in row) + "\n"
        return text
    else:
        return "Unsupported file type"
